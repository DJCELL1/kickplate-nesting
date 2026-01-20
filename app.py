import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from typing import List, Dict, Tuple, Optional, Any
import re
from dataclasses import dataclass, asdict
import io
import pdfplumber
import openpyxl
from datetime import datetime

# PDF Generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image, Frame, PageTemplate
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# Visualization
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from matplotlib.backends.backend_agg import FigureCanvasAgg

# ========== DATA CLASSES ==========
@dataclass
class Piece:
    code: str
    description: str
    width: int
    height: int
    material: str
    qty: int

@dataclass
class Placement:
    x: int
    y: int
    width: int
    height: int
    code: str
    description: str
    rotated: bool

@dataclass
class Sheet:
    id: int
    placements: List[Placement]
    waste_area: float

@dataclass
class DoorLabel:
    door_number: str
    area: str
    kickplate_code: str
    width: int
    height: int
    project_code: str
    project_name: str
    material: str
    quantity: int = 1
    
    def label_text(self, include_project: bool = True) -> str:
        text = f"DOOR: {self.door_number}\n"
        text += f"AREA: {self.area}\n"
        text += f"KICKPLATE: {self.width}×{self.height}mm\n"
        if include_project:
            text += f"PROJECT: {self.project_code}"
        return text
    
    def to_dict(self) -> dict:
        return asdict(self)

# ========== INTEGRATION FUNCTIONS ==========
def extract_pieces_from_labels(labels: List[DoorLabel]) -> List[Piece]:
    """Convert DoorLabels to Pieces for nesting optimizer"""
    # Group by kickplate size and material
    piece_counts = {}
    
    for label in labels:
        key = (label.width, label.height, label.material)
        if key not in piece_counts:
            piece_counts[key] = {
                'count': 0,
                'door_numbers': [],
                'areas': []
            }
        piece_counts[key]['count'] += 1
        piece_counts[key]['door_numbers'].append(label.door_number)
        piece_counts[key]['areas'].append(label.area)
    
    # Convert to Pieces
    pieces = []
    for (width, height, material), data in piece_counts.items():
        # Create a description with sample door numbers
        sample_doors = ', '.join(data['door_numbers'][:3])  # Show first 3 doors
        if len(data['door_numbers']) > 3:
            sample_doors += f" and {len(data['door_numbers']) - 3} more"
        
        code = f"KP{width}{height}{material}"
        description = f"{width}×{height}mm {material} - Doors: {sample_doors}"
        
        pieces.append(Piece(
            code=code,
            description=description,
            width=width,
            height=height,
            material=material,
            qty=data['count']
        ))
    
    return pieces

# ========== NESTING OPTIMIZER ==========
class KickplateNester:
    def __init__(self, stock_width: int, stock_height: int, kerf_width: int, grain_direction: str):
        self.stock_width = stock_width
        self.stock_height = stock_height
        self.kerf_width = kerf_width
        self.grain_direction = grain_direction
    
    def parse_kickplate_code(self, code: str) -> Optional[Dict]:
        match = re.match(r'^KP(\d{3,4})(\d{3})(.+)$', code, re.IGNORECASE)
        if not match:
            return None
        
        width = int(match.group(1))
        height = int(match.group(2))
        material = match.group(3)
        
        if width <= 0 or height <= 0:
            return None
        
        return {'width': width, 'height': height, 'material': material}
    
    def find_gaps(self, placements: List[Placement], max_w: int, max_h: int, kerf: int) -> List[Dict]:
        gaps = []
        sorted_placements = sorted(placements, key=lambda p: (p.y, p.x))
        
        for p in sorted_placements:
            gaps.append({
                'x': p.x + p.width + kerf,
                'y': p.y,
                'width': max_w - (p.x + p.width + kerf),
                'height': p.height
            })
            
            gaps.append({
                'x': p.x,
                'y': p.y + p.height + kerf,
                'width': p.width,
                'height': max_h - (p.y + p.height + kerf)
            })
        
        valid_gaps = []
        for gap in gaps:
            if gap['width'] <= 0 or gap['height'] <= 0:
                continue
            if gap['x'] >= max_w or gap['y'] >= max_h:
                continue
            
            overlaps = False
            for p in placements:
                if not (gap['x'] >= p.x + p.width + kerf or
                       gap['x'] + gap['width'] <= p.x or
                       gap['y'] >= p.y + p.height + kerf or
                       gap['y'] + gap['height'] <= p.y):
                    overlaps = True
                    break
            
            if not overlaps:
                valid_gaps.append(gap)
        
        return sorted(valid_gaps, key=lambda g: g['x'] + g['y'])
    
    def find_placement(self, sheet: Sheet, piece: Piece) -> Optional[Placement]:
        orientations = []
        if self.grain_direction == 'none':
            orientations = [
                {'w': piece.width, 'h': piece.height, 'rotated': False},
                {'w': piece.height, 'h': piece.width, 'rotated': True}
            ]
        else:
            orientations = [{'w': piece.width, 'h': piece.height, 'rotated': False}]
        
        for orient in orientations:
            if len(sheet.placements) == 0:
                if orient['w'] <= self.stock_width and orient['h'] <= self.stock_height:
                    return Placement(
                        x=0, y=0,
                        width=orient['w'],
                        height=orient['h'],
                        code=piece.code,
                        description=piece.description,
                        rotated=orient['rotated']
                    )
            else:
                gaps = self.find_gaps(sheet.placements, self.stock_width, 
                                     self.stock_height, self.kerf_width)
                for gap in gaps:
                    if orient['w'] <= gap['width'] and orient['h'] <= gap['height']:
                        return Placement(
                            x=gap['x'], y=gap['y'],
                            width=orient['w'],
                            height=orient['h'],
                            code=piece.code,
                            description=piece.description,
                            rotated=orient['rotated']
                        )
        
        return None
    
    def nest_pieces(self, pieces: List[Piece]) -> List[Sheet]:
        sheets = []
        current_sheet = Sheet(id=0, placements=[], 
                            waste_area=self.stock_width * self.stock_height)
        sheets.append(current_sheet)
        
        all_pieces = []
        for piece in pieces:
            for i in range(piece.qty):
                all_pieces.append(piece)
        
        all_pieces.sort(key=lambda p: p.width * p.height, reverse=True)
        
        for piece in all_pieces:
            placed = False
            
            for sheet in sheets:
                placement = self.find_placement(sheet, piece)
                if placement:
                    sheet.placements.append(placement)
                    used_area = sum(p.width * p.height for p in sheet.placements)
                    sheet.waste_area = (self.stock_width * self.stock_height) - used_area
                    placed = True
                    break
            
            if not placed:
                current_sheet = Sheet(id=len(sheets), placements=[], 
                                    waste_area=self.stock_width * self.stock_height)
                sheets.append(current_sheet)
                placement = self.find_placement(current_sheet, piece)
                if placement:
                    current_sheet.placements.append(placement)
                    used_area = sum(p.width * p.height for p in current_sheet.placements)
                    current_sheet.waste_area = (self.stock_width * self.stock_height) - used_area
        
        return sheets

# ========== DOOR LABEL PARSERS ==========
class DoorLabelParser:
    """Unified parser for all door label formats - SKIPS DOORS WITHOUT KICKPLATE DIMENSIONS"""
    
    def __init__(self):
        self.kickplate_pattern = r'KP(\d{3,4})(\d{3})([A-Z]+)'
        self.project_pattern = r'([Qq]\d+[A-Za-z]*)\s*:\s*(.+?)(?:\n|$)'
    
    def auto_parse(self, file, file_type: str = None, project_info: Dict = None) -> List[DoorLabel]:
        """Auto-detect format and parse"""
        if file_type is None:
            file_type = self._detect_file_type(file)
        
        if file_type == 'pdf':
            return self.parse_pdf_fixed(file, project_info)
        elif file_type == 'csv':
            return self.parse_csv(file, project_info)
        elif file_type == 'excel':
            return self.parse_excel(file, project_info)
        else:
            raise ValueError(f"Unsupported file type: {file_type}")
    
    def _detect_file_type(self, file) -> str:
        """Detect file type from name or content"""
        name = file.name.lower() if hasattr(file, 'name') else str(file).lower()
        
        if name.endswith('.pdf'):
            return 'pdf'
        elif name.endswith('.csv'):
            return 'csv'
        elif name.endswith(('.xlsx', '.xls')):
            return 'excel'
        else:
            # Try to infer from content
            content = file.read(100)
            file.seek(0)
            
            if b'%PDF' in content:
                return 'pdf'
            elif b',' in content or b';' in content:
                return 'csv'
            else:
                return 'unknown'
    
    
    def parse_pdf_fixed(self, pdf_file, project_info: Dict = None) -> List[DoorLabel]:
        """Parse Hardware Direct PDF - SKIPS doors without kickplate dimensions"""
        labels = []
        
        # Save original position for later reading
        original_position = pdf_file.tell() if hasattr(pdf_file, 'tell') else None
        
        # If no project info provided, extract it from PDF
        if project_info is None:
            project_info = self._extract_project_info_from_pdf(pdf_file)
            print(f"DEBUG: Auto-extracted project info: {project_info['project_code']}")
        
        # Reset file position if we moved it
        if original_position is not None:
            pdf_file.seek(original_position)
        
        # Read all text from PDF
        with pdfplumber.open(pdf_file) as pdf:
            all_text = ""
            for page in pdf.pages:
                text = page.extract_text()
                all_text += text + "\n"
        
        # Clean and normalize text
        all_text = all_text.replace('\r', '\n')
        
        print("=" * 80)
        print("DEBUG: START OF PDF PARSING")
        print(f"DEBUG: Project: {project_info['project_code']} - {project_info['project_name']}")
        print("=" * 80)
        
        # Split into lines
        lines = all_text.split('\n')
        
        # Process lines
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            
            # Skip empty lines and headers
            if (not line or 
                'Code Description' in line or 
                'Door Area' in line or 
                'Page' in line or
                'ProMaster' in line or
                'Hardware Direct' in line or
                'Phone' in line or
                'Consultant:' in line or
                'Finishing instructions' in line or
                'All doors' in line):
                i += 1
                continue
            
            # Check if this is a kickplate line (starts with KP)
            if line.startswith('KP'):
                print(f"\nDEBUG [{i}]: Found KP line: {line}")
                
                # Check if this kickplate has ACTUAL dimensions
                # Look for "W XXX H XXX" pattern (with numbers)
                has_actual_dimensions = bool(re.search(r'W\s+\d{3,4}\s+H\s+\d{3}', line))
                
                if not has_actual_dimensions:
                    # Check if it's explicitly missing dimensions: "W H T" or "W H"
                    if 'W H' in line or 'W H T' in line:
                        print(f"DEBUG: SKIPPING - No dimensions found: {line}")
                        # Skip this kickplate AND all its associated doors
                        j = i + 1
                        while j < len(lines) and not lines[j].strip().startswith('KP'):
                            door_line = lines[j].strip()
                            if door_line and re.match(r'^[DWM]\.', door_line):
                                print(f"DEBUG: Also skipping door (no kickplate): {door_line}")
                            j += 1
                        i = j  # Skip to next kickplate
                        continue
                
                # Extract kickplate info from this line
                kp_info = self._extract_kickplate_from_debug_line(line)
                
                if kp_info and kp_info.get('has_dimensions', False):
                    print(f"DEBUG: Extracted KP with dimensions: {kp_info['width']}×{kp_info['height']}")
                    
                    # Now look for door lines that follow
                    j = i + 1
                    door_count = 0
                    while j < len(lines) and not lines[j].strip().startswith('KP'):
                        door_line = lines[j].strip()
                        if door_line and re.match(r'^[DWM]\.', door_line):
                            # Extract door info
                            door_info = self._extract_door_from_debug_line(door_line)
                            if door_info:
                                # Create label ONLY if we have valid dimensions
                                label = DoorLabel(
                                    door_number=door_info['door'],
                                    area=door_info['area'],
                                    kickplate_code=f"KP{str(kp_info['width']).zfill(4)}{str(kp_info['height']).zfill(3)}{kp_info['material']}",
                                    width=kp_info['width'],
                                    height=kp_info['height'],
                                    project_code=project_info['project_code'],
                                    project_name=project_info['project_name'],
                                    material=kp_info['material'],
                                    quantity=1
                                )
                                labels.append(label)
                                door_count += 1
                                print(f"DEBUG [{j}]: Created label - Door: {door_info['door']}, Size: {kp_info['width']}×{kp_info['height']}")
                        j += 1
                    
                    print(f"DEBUG: Found {door_count} doors for this kickplate")
                    
                    # Skip to after the door lines
                    i = j
                else:
                    # No valid dimensions, skip to next kickplate
                    print(f"DEBUG: Skipping - no valid dimensions extracted")
                    j = i + 1
                    while j < len(lines) and not lines[j].strip().startswith('KP'):
                        j += 1
                    i = j
            else:
                i += 1
        
        print(f"\nDEBUG: Total labels parsed: {len(labels)}")
        print("=" * 80)
        print("DEBUG: END OF PDF PARSING")
        print("=" * 80)
        
        return labels
    
    def _extract_kickplate_from_debug_line(self, line: str) -> Optional[Dict]:
        """Extract kickplate info - returns None if no dimensions found"""
        # Example: "KP600300SSS Kickplate W 885 H 300 T 1"  ← HAS dimensions
        # Example: "KP860300SSS KICKPLATE 860X300.1.2 W H T 2"  ← NO dimensions!
        
        # Find kickplate code
        kp_match = re.search(r'(KP\d{3,4}\d{3}[A-Z]+)', line)
        if not kp_match:
            return None
        
        code = kp_match.group(1)
        
        # Parse code for material
        code_match = re.match(self.kickplate_pattern, code)
        if not code_match:
            return None
        
        material = code_match.group(3)
        
        # Check if this line has ACTUAL dimensions (W XXX H XXX pattern)
        # If it's missing dimensions like "W H T", return None
        has_dimensions_pattern = r'W\s+(\d{3,4})\s+H\s+(\d{3})'
        dim_match = re.search(has_dimensions_pattern, line)
        
        if not dim_match:
            # Also check for patterns like "W H" or "W H T" (empty dimensions)
            if re.search(r'W\s+H', line) or 'W H T' in line:
                return None  # SKIP THIS KICKPLATE - NO DIMENSIONS
        
        # Extract ACTUAL width and height from "W 885 H 300" pattern
        dim_match = re.search(r'W\s+(\d{3,4})\s+H\s+(\d{3})', line)
        if dim_match:
            width = int(dim_match.group(1))
            height = int(dim_match.group(2))
        else:
            # Try to find any numbers that could be dimensions
            numbers = re.findall(r'\b(\d{3,4})\b', line)
            
            # Filter out code numbers
            code_numbers = re.findall(r'\d+', code)
            possible_dims = []
            
            for num in numbers:
                num_int = int(num)
                # Skip numbers that are in the code or not reasonable kickplate dimensions
                if num not in code_numbers and 200 <= num_int <= 2000:
                    possible_dims.append(num_int)
            
            if len(possible_dims) >= 2:
                # Sort to get width and height (width is usually larger)
                possible_dims.sort(reverse=True)
                width = possible_dims[0]
                height = possible_dims[1] if possible_dims[1] <= 400 else 300
            else:
                # Not enough dimension numbers found
                return None  # SKIP THIS KICKPLATE
        
        return {
            'code': code,
            'width': width,
            'height': height,
            'material': material,
            'has_dimensions': True
        }
    
    def _extract_door_from_debug_line(self, line: str) -> Optional[Dict]:
        """Extract door info from debug format line"""
        # Example: "D.26-K PE Office 2"
        # Example: "W.15 WC 6 1"
        
        # Remove extra spaces
        line = re.sub(r'\s+', ' ', line.strip())
        
        # Split by spaces
        parts = line.split()
        if len(parts) < 2:
            return None
        
        # First part is door number
        door = parts[0]
        
        # Validate door format
        if not re.match(r'^[DWM]\.\d', door):
            return None
        
        # The rest is area + optional sequence number
        area_parts = parts[1:]
        
        # Remove sequence number if it's the last part and is a single digit
        if area_parts and re.match(r'^\d+$', area_parts[-1]):
            area_parts = area_parts[:-1]
        
        area = ' '.join(area_parts)
        
        return {'door': door, 'area': area}
    
    def _extract_project_info_from_pdf(self, pdf_file) -> Dict:
        """Extract project info from PDF - ENHANCED VERSION"""
        try:
            with pdfplumber.open(pdf_file) as pdf:
                first_page = pdf.pages[0]
                text = first_page.extract_text()
                
                print(f"DEBUG: Extracting project info from text:\n{text[:500]}")
                
                # Pattern 1: "Q30683B: Matamata Indoor Sports Facility"
                match = re.search(r'([Qq]\d+[A-Za-z]*)\s*:\s*(.+?)(?:\n|$)', text)
                if match:
                    project_code = match.group(1).strip()
                    project_name = match.group(2).strip()
                    print(f"DEBUG: Found project: {project_code} - {project_name}")
                    return {
                        'project_code': project_code,
                        'project_name': project_name
                    }
                
                # Pattern 2: Look for project number in various formats
                project_patterns = [
                    r'Project\s*[#:]?\s*([Qq]?\d+[A-Za-z]*)',  # Project: Q30683B
                    r'Job\s*[#:]?\s*([Qq]?\d+[A-Za-z]*)',      # Job: Q30683B
                    r'Ref\s*[#:]?\s*([Qq]?\d+[A-Za-z]*)',      # Ref: Q30683B
                    r'([Qq]\d+[A-Za-z]*)\s+[-:]',              # Q30683B -
                    r'\b([Qq]\d+[A-Za-z]*)\b',                 # Just Q30683B
                ]
                
                for pattern in project_patterns:
                    match = re.search(pattern, text)
                    if match:
                        project_code = match.group(1).strip()
                        print(f"DEBUG: Found project code: {project_code}")
                        
                        # Try to find project name (text after code)
                        name_match = re.search(f'{re.escape(project_code)}[\\s:]+(.+?)(?:\\n|$)', text)
                        if name_match:
                            project_name = name_match.group(1).strip()
                        else:
                            project_name = f"Project {project_code}"
                        
                        return {
                            'project_code': project_code,
                            'project_name': project_name
                        }
                
                # Pattern 3: Look for any Q-number in the first few lines
                lines = text.split('\n')
                for line in lines[:10]:  # Check first 10 lines
                    match = re.search(r'([Qq]\d+[A-Za-z]*)', line)
                    if match:
                        project_code = match.group(1).strip()
                        print(f"DEBUG: Found project in line: {line}")
                        
                        # Extract name from same line after code
                        parts = line.split(project_code)
                        if len(parts) > 1 and parts[1].strip():
                            project_name = parts[1].split('\n')[0].strip(': -')
                        else:
                            project_name = f"Project {project_code}"
                        
                        return {
                            'project_code': project_code,
                            'project_name': project_name
                        }
                
                print("DEBUG: No project info found, using defaults")
                return {
                    'project_code': 'UNKNOWN',
                    'project_name': 'Unknown Project'
                }
                
        except Exception as e:
            print(f"ERROR extracting project info: {e}")
            return {
                'project_code': 'UNKNOWN',
                'project_name': 'Unknown Project'
            }
    
    def parse_csv(self, csv_file, project_info: Dict = None) -> List[DoorLabel]:
        """Parse order CSV files"""
        df = pd.read_csv(csv_file)
        labels = []
        
        if project_info is None:
            project_info = {
                'project_code': 'UNKNOWN',
                'project_name': 'Unknown Project'
            }
        
        for _, row in df.iterrows():
            part_code = str(row.get('PartCode', ''))
            
            if part_code.startswith('KP'):
                match = re.match(self.kickplate_pattern, part_code)
                if match:
                    width = int(match.group(1))
                    height = int(match.group(2))
                    material = match.group(3)
                    qty = int(row.get('ProductQuantity', 1))
                    
                    for i in range(qty):
                        labels.append(DoorLabel(
                            door_number=f"TBD-{len(labels)+1}",
                            area="To be assigned",
                            kickplate_code=part_code,
                            width=width,
                            height=height,
                            project_code=project_info['project_code'],
                            project_name=project_info['project_name'],
                            material=material,
                            quantity=1
                        ))
        
        return labels
    
    def parse_excel(self, excel_file, project_info: Dict = None) -> List[DoorLabel]:
        """Parse Excel door schedules"""
        df = pd.read_excel(excel_file)
        labels = []
        
        if project_info is None:
            project_info = {
                'project_code': 'UNKNOWN',
                'project_name': 'Unknown Project'
            }
        
        # Try to identify columns
        door_col = None
        area_col = None
        size_col = None
        code_col = None
        
        for col in df.columns:
            col_lower = str(col).lower()
            if any(word in col_lower for word in ['door', 'number', 'ref']):
                door_col = col
            elif any(word in col_lower for word in ['area', 'location', 'room']):
                area_col = col
            elif any(word in col_lower for word in ['size', 'dimension']):
                size_col = col
            elif any(word in col_lower for word in ['code', 'part']):
                code_col = col
        
        for _, row in df.iterrows():
            # Try to extract from kickplate code
            kickplate_code = None
            if code_col and pd.notna(row[code_col]):
                kickplate_code = str(row[code_col])
            elif size_col and pd.notna(row[size_col]):
                # Try to construct code from size
                size_str = str(row[size_col])
                if '×' in size_str or 'x' in size_str:
                    parts = re.split('[×x]', size_str.replace('mm', '').strip())
                    if len(parts) == 2:
                        kickplate_code = f"KP{parts[0].zfill(3)}{parts[1].zfill(3)}SSS"
            
            if kickplate_code and kickplate_code.startswith('KP'):
                match = re.match(self.kickplate_pattern, kickplate_code)
                if match:
                    width = int(match.group(1))
                    height = int(match.group(2))
                    material = match.group(3)
                    
                    door_number = str(row[door_col]) if door_col and pd.notna(row[door_col]) else f"TBD-{len(labels)+1}"
                    area = str(row[area_col]) if area_col and pd.notna(row[area_col]) else "To be assigned"
                    
                    labels.append(DoorLabel(
                        door_number=door_number,
                        area=area,
                        kickplate_code=kickplate_code,
                        width=width,
                        height=height,
                        project_code=project_info['project_code'],
                        project_name=project_info['project_name'],
                        material=material,
                        quantity=1
                    ))
        
        return labels
    
    def _deduplicate_labels(self, labels: List[DoorLabel]) -> List[DoorLabel]:
        """Remove duplicate labels"""
        seen = set()
        unique = []
        
        for label in labels:
            key = (label.door_number, label.kickplate_code, label.width, label.height)
            if key not in seen:
                seen.add(key)
                unique.append(label)
        
        return unique

# ========== VISUALIZATION FUNCTIONS ==========
def create_sheet_visualization(sheet: Sheet, stock_width: int, stock_height: int, 
                               grain_direction: str) -> go.Figure:
    fig = go.Figure()
    
    fig.add_shape(
        type="rect",
        x0=0, y0=0, x1=stock_width, y1=stock_height,
        line=dict(color="black", width=3),
        fillcolor="white"
    )
    
    for i in range(100, stock_width, 100):
        fig.add_shape(type="line", x0=i, y0=0, x1=i, y1=stock_height,
                     line=dict(color="lightgray", width=1, dash="dot"))
    for i in range(100, stock_height, 100):
        fig.add_shape(type="line", x0=0, y0=i, x1=stock_width, y1=i,
                     line=dict(color="lightgray", width=1, dash="dot"))
    
    colors_list = ['#3b82f6', '#10b981', '#f59e0b', '#ef4444', '#8b5cf6', '#ec4899']
    for i, p in enumerate(sheet.placements):
        color = colors_list[i % len(colors_list)]
        
        fig.add_shape(
            type="rect",
            x0=p.x, y0=p.y, x1=p.x + p.width, y1=p.y + p.height,
            line=dict(color="darkblue", width=2),
            fillcolor=color,
            opacity=0.7
        )
        
        fig.add_annotation(
            x=p.x + p.width/2,
            y=p.y + p.height/2,
            text=f"{p.code}<br>{p.width}×{p.height}{'↻' if p.rotated else ''}",
            showarrow=False,
            font=dict(color="white", size=10, family="monospace"),
            bgcolor=color,
            opacity=0.9
        )
    
    efficiency = 100 - (sheet.waste_area / (stock_width * stock_height) * 100)
    fig.update_layout(
        title=f"Sheet {sheet.id + 1} - Efficiency: {efficiency:.1f}%",
        xaxis=dict(title="Width (mm)", range=[0, stock_width]),
        yaxis=dict(title="Height (mm)", range=[0, stock_height], scaleanchor="x", scaleratio=1),
        width=900,
        height=600,
        showlegend=False
    )
    
    return fig

def create_matplotlib_diagram(sheet: Sheet, stock_width: int, stock_height: int, 
                             grain_direction: str) -> io.BytesIO:
    fig, ax = plt.subplots(figsize=(12, 8))
    
    ax.set_xlim(0, stock_width)
    ax.set_ylim(0, stock_height)
    ax.set_aspect('equal')
    
    for i in range(0, stock_width + 1, 100):
        ax.axvline(i, color='lightgray', linewidth=0.5, linestyle='--', alpha=0.5)
    for i in range(0, stock_height + 1, 100):
        ax.axhline(i, color='lightgray', linewidth=0.5, linestyle='--', alpha=0.5)
    
    boundary = patches.Rectangle((0, 0), stock_width, stock_height, 
                                linewidth=3, edgecolor='black', 
                                facecolor='white', fill=True)
    ax.add_patch(boundary)
    
    color_palette = ['#3b82f6', '#10b981', '#f59e0b', '#ef4444', '#8b5cf6', '#ec4899']
    
    for i, p in enumerate(sheet.placements):
        color = color_palette[i % len(color_palette)]
        
        rect = patches.Rectangle((p.x, p.y), p.width, p.height,
                                linewidth=2, edgecolor='darkblue',
                                facecolor=color, alpha=0.7)
        ax.add_patch(rect)
        
        text_x = p.x + p.width / 2
        text_y = p.y + p.height / 2
        
        label = f"{p.code}\n{p.width}×{p.height}"
        if p.rotated:
            label += " ↻"
        
        ax.text(text_x, text_y, label,
               ha='center', va='center',
               fontsize=8, fontweight='bold',
               color='white',
               bbox=dict(boxstyle='round,pad=0.3', facecolor=color, alpha=0.9))
    
    efficiency = 100 - (sheet.waste_area / (stock_width * stock_height) * 100)
    ax.set_xlabel('Width (mm)', fontsize=10)
    ax.set_ylabel('Height (mm)', fontsize=10)
    ax.set_title(f'Sheet {sheet.id + 1} - Efficiency: {efficiency:.1f}%', 
                fontsize=12, fontweight='bold')
    
    ax.grid(True, alpha=0.3)
    
    buf = io.BytesIO()
    plt.tight_layout()
    plt.savefig(buf, format='png', dpi=150, bbox_inches='tight')
    plt.close(fig)
    buf.seek(0)
    
    return buf

# ========== PDF GENERATION ==========
def generate_pdf_cutlist(sheets: List[Sheet], stock_width: int, stock_height: int, 
                        grain_direction: str, reference_number: str = None, 
                        filename: str = "cut_list.pdf") -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                           rightMargin=20, leftMargin=20,
                           topMargin=30, bottomMargin=20)
    
    elements = []
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=15,
        alignment=TA_CENTER
    )
    
    title_text = "Kickplate Cutting List"
    if reference_number:
        title_text += f" - {reference_number}"
    title = Paragraph(title_text, title_style)
    elements.append(title)
    
    info_style = ParagraphStyle(
        'Info',
        parent=styles['Normal'],
        fontSize=10,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    
    info_text = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}<br/>"
    info_text += f"Stock Size: {stock_width}mm × {stock_height}mm<br/>"
    info_text += f"Grain Direction: {grain_direction.title()}<br/>"
    info_text += f"Total Sheets: {len(sheets)}"
    
    elements.append(Paragraph(info_text, info_style))
    elements.append(Spacer(1, 20))
    
    summary_data = [['Sheet', 'Pieces', 'Efficiency', 'Waste Area']]
    for sheet in sheets:
        efficiency = 100 - (sheet.waste_area / (stock_width * stock_height) * 100)
        summary_data.append([
            f"Sheet {sheet.id + 1}",
            str(len(sheet.placements)),
            f"{efficiency:.1f}%",
            f"{sheet.waste_area:.0f} mm²"
        ])
    
    summary_table = Table(summary_data, colWidths=[60*mm, 40*mm, 40*mm, 50*mm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
    ]))
    
    elements.append(summary_table)
    elements.append(PageBreak())
    
    for sheet in sheets:
        header_text = f"<b>Sheet {sheet.id + 1}</b> - {len(sheet.placements)} pieces"
        if reference_number:
            header_text += f" | Ref: {reference_number}"
        sheet_title = Paragraph(header_text, styles['Heading2'])
        elements.append(sheet_title)
        elements.append(Spacer(1, 10))
        
        piece_counts = {}
        for p in sheet.placements:
            key = f"{p.width}×{p.height}mm"
            piece_counts[key] = piece_counts.get(key, 0) + 1
        
        cutting_list_data = [['☐', 'Size', 'Qty']]
        for size, qty in sorted(piece_counts.items()):
            cutting_list_data.append(['☐', size, str(qty)])
        
        cutting_list_table = Table(cutting_list_data, colWidths=[8*mm, 30*mm, 12*mm])
        cutting_list_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
            ('ALIGN', (2, 0), (2, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        img_buffer = create_matplotlib_diagram(sheet, stock_width, stock_height, grain_direction)
        diagram_img = Image(img_buffer, width=210*mm, height=140*mm)
        
        layout_table = Table([[diagram_img, cutting_list_table]], colWidths=[210*mm, 50*mm])
        layout_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'LEFT'),
        ]))
        
        elements.append(layout_table)
        
        if sheet.id < len(sheets) - 1:
            elements.append(PageBreak())
        else:
            elements.append(Spacer(1, 20))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()

def generate_label_pdf(labels: List[DoorLabel], label_size: tuple = (90, 90), 
                      labels_per_page: tuple = (3, 8), include_project: bool = True,
                      page_size: tuple = letter) -> bytes:
    """
    Generate PDF with door labels
    label_size: (width, height) in mm
    labels_per_page: (cols, rows)
    """
    buffer = io.BytesIO()
    
    # Convert mm to points (1 mm = 2.83465 points)
    label_width_pt = label_size[0] * 2.83465
    label_height_pt = label_size[1] * 2.83465
    
    doc = SimpleDocTemplate(buffer, pagesize=page_size,
                           leftMargin=10, rightMargin=10,
                           topMargin=10, bottomMargin=10)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'LabelTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#1e40af'),
        alignment=TA_CENTER,
        spaceAfter=20
    )
    
    project_code = labels[0].project_code if labels else "UNKNOWN"
    title = Paragraph(f"Door Labels - Project {project_code}", title_style)
    elements.append(title)
    
    # Create label grid
    label_style = ParagraphStyle(
        'LabelText',
        parent=styles['Normal'],
        fontSize=10,
        leading=12,
        fontName='Helvetica-Bold',
        alignment=TA_CENTER,
        textColor=colors.black,
        borderWidth=1,
        borderColor=colors.black,
        borderRadius=5,
        backColor=colors.white,
        padding=(6, 6, 6, 6)
    )
    
    # Group labels for grid layout
    cols, rows = labels_per_page
    labels_per_sheet = cols * rows
    
    for i in range(0, len(labels), labels_per_sheet):
        page_labels = labels[i:i + labels_per_sheet]
        
        # Create grid data
        grid_data = []
        for row in range(rows):
            row_data = []
            for col in range(cols):
                idx = row * cols + col
                if idx < len(page_labels):
                    label = page_labels[idx]
                    text = label.label_text(include_project=include_project)
                    row_data.append(Paragraph(text, label_style))
                else:
                    row_data.append('')
            grid_data.append(row_data)
        
        # Create table for labels
        label_table = Table(grid_data, 
                           colWidths=[label_width_pt] * cols,
                           rowHeights=[label_height_pt] * rows)
        
        label_table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 1, colors.gray),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BACKGROUND', (0, 0), (-1, -1), colors.white),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ]))
        
        elements.append(label_table)
        
        # Add page break if not last page
        if i + labels_per_sheet < len(labels):
            elements.append(PageBreak())
    
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()

# ========== STREAMLIT APP ==========
def main():
    st.set_page_config(page_title="Kickplate Nesting Optimizer", layout="wide")

    # Header with logo
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        import os
        logo_path = "Logos-01.jpg"
        if os.path.exists(logo_path):
            st.image(logo_path, use_container_width=True)
        else:
            st.markdown("<h1 style='text-align: center; color: #F47920;'>HDL</h1>", unsafe_allow_html=True)

    st.markdown("<h2 style='text-align: center; color: #2B2B2B;'>Kickplate Nesting Optimizer</h2>", unsafe_allow_html=True)
    st.markdown("<p style='text-align: center; color: #666;'>Optimize cutting layouts and generate door labels for kickplates</p>", unsafe_allow_html=True)
    
    # Initialize session state
    if 'manual_items' not in st.session_state:
        st.session_state.manual_items = []
    if 'manual_sheets' not in st.session_state:
        st.session_state.manual_sheets = None
    if 'door_labels' not in st.session_state:
        st.session_state.door_labels = []
    if 'project_info' not in st.session_state:
        st.session_state.project_info = {'project_code': '', 'project_name': ''}
    if 'label_pieces' not in st.session_state:
        st.session_state.label_pieces = None
    if 'label_sheets' not in st.session_state:
        st.session_state.label_sheets = None
    if 'parsed_project_info' not in st.session_state:
        st.session_state.parsed_project_info = None
    
    # Sidebar configuration
    with st.sidebar:
        st.header("⚙️ Configuration")
        
        stock_width = st.number_input("Stock Width (mm)", value=2400, min_value=100, step=10)
        stock_height = st.number_input("Stock Height (mm)", value=1200, min_value=100, step=10)
        kerf_width = st.number_input("Kerf Width (mm)", value=0, min_value=0, step=1)
        grain_direction = st.selectbox(
            "Grain Direction",
            options=['horizontal', 'vertical', 'none'],
            index=0,
            format_func=lambda x: {
                'horizontal': 'Horizontal (no rotation)',
                'vertical': 'Vertical (no rotation)',
                'none': 'No preference (allow rotation)'
            }[x]
        )
        
        st.divider()
        st.header("🏷️ Label Settings")
        label_width = st.number_input("Label Width (mm)", value=60, min_value=50, max_value=200)
        label_height = st.number_input("Label Height (mm)", value=70, min_value=30, max_value=150)
        include_project = st.checkbox("Include Project Code on Labels", value=True)
        
        st.divider()
        st.header("📋 Project Info")
        project_code = st.text_input("Project Code", value=st.session_state.project_info['project_code'])
        project_name = st.text_input("Project Name", value=st.session_state.project_info['project_name'])
        
        if project_code or project_name:
            st.session_state.project_info = {
                'project_code': project_code,
                'project_name': project_name
            }
    
    # Main content with tabs
    tab1, tab2, tab3, tab4 = st.tabs(["📤 Upload Order CSV", "➕ Manual Entry", "🏷️ Door Labels", "🔧 Debug Tools"])
    
    # ========== TAB 1: Upload Order CSV ==========
    with tab1:
        st.subheader("Upload Order CSV")
        uploaded_file = st.file_uploader("Choose a CSV file", type=['csv'], key="csv_uploader")
        
        if uploaded_file:
            df = pd.read_csv(uploaded_file)
            
            reference_number = uploaded_file.name.replace('_ShipmentProductWithCostsAndPrice.csv', '').replace('.csv', '')
            
            st.success(f"Loaded {len(df)} rows from {uploaded_file.name}")
            st.info(f"📋 Reference: **{reference_number}**")
            
            kickplate_df = df[df['PartCode'].str.startswith('KP', na=False)].copy()
            
            if len(kickplate_df) > 0:
                st.write(f"Found {len(kickplate_df)} kickplate items:")
                st.dataframe(kickplate_df[['PartCode', 'Description', 'ProductQuantity', 'ProductPrice']])
                
                col1, col2, col3, col4 = st.columns(4)
                with col1:
                    st.metric("Unique Items", len(kickplate_df))
                with col2:
                    total_qty = kickplate_df['ProductQuantity'].sum()
                    st.metric("Total Pieces", int(total_qty))
                with col3:
                    total_cost = (kickplate_df['ProductCost'] * kickplate_df['ProductQuantity']).sum()
                    st.metric("Total Cost", f"${total_cost:.2f}")
                with col4:
                    total_price = (kickplate_df['ProductPrice'] * kickplate_df['ProductQuantity']).sum()
                    st.metric("Total Revenue", f"${total_price:.2f}")
                
                if st.button("🚀 Generate Cut List", type="primary", use_container_width=True, key="csv_generate"):
                    nester = KickplateNester(stock_width, stock_height, kerf_width, grain_direction)
                    
                    pieces = []
                    for _, row in kickplate_df.iterrows():
                        parsed = nester.parse_kickplate_code(row['PartCode'])
                        if parsed:
                            pieces.append(Piece(
                                code=row['PartCode'],
                                description=row['Description'],
                                width=parsed['width'],
                                height=parsed['height'],
                                material=parsed['material'],
                                qty=int(row['ProductQuantity'])
                            ))
                    
                    with st.spinner("Optimizing layout..."):
                        sheets = nester.nest_pieces(pieces)
                    
                    st.success(f"✅ Generated cut list with {len(sheets)} sheet(s)")
                    
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Sheets Required", len(sheets))
                    with col2:
                        total_pieces = sum(len(s.placements) for s in sheets)
                        st.metric("Total Pieces", total_pieces)
                    with col3:
                        avg_efficiency = sum(100 - (s.waste_area / (stock_width * stock_height) * 100) for s in sheets) / len(sheets)
                        st.metric("Avg Efficiency", f"{avg_efficiency:.1f}%")
                    
                    for sheet in sheets:
                        with st.expander(f"📋 Sheet {sheet.id + 1} - {len(sheet.placements)} pieces", expanded=True):
                            fig = create_sheet_visualization(sheet, stock_width, stock_height, grain_direction)
                            st.plotly_chart(fig, use_container_width=True)
                            
                            st.subheader("Cutting Instructions")
                            cut_data = []
                            for i, p in enumerate(sheet.placements):
                                cut_data.append({
                                    'Piece #': i + 1,
                                    'Part Code': p.code,
                                    'Description': p.description,
                                    'X (mm)': p.x,
                                    'Y (mm)': p.y,
                                    'Width (mm)': p.width,
                                    'Height (mm)': p.height,
                                    'Rotated': '↻ Yes' if p.rotated else 'No'
                                })
                            st.dataframe(pd.DataFrame(cut_data), use_container_width=True)
                    
                    cut_list_data = []
                    for sheet in sheets:
                        for i, p in enumerate(sheet.placements):
                            cut_list_data.append({
                                'Sheet': sheet.id + 1,
                                'Piece #': i + 1,
                                'Part Code': p.code,
                                'Description': p.description,
                                'X Position': p.x,
                                'Y Position': p.y,
                                'Width': p.width,
                                'Height': p.height,
                                'Rotated': 'Yes' if p.rotated else 'No'
                            })
                    
                    csv_data = pd.DataFrame(cut_list_data).to_csv(index=False)
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        st.download_button(
                            label="📥 Download CSV",
                            data=csv_data,
                            file_name="cut_list.csv",
                            mime="text/csv",
                            use_container_width=True
                        )
                    with col2:
                        pdf_data = generate_pdf_cutlist(sheets, stock_width, stock_height, grain_direction, 
                                                       reference_number=reference_number)
                        st.download_button(
                            label="📄 Download PDF",
                            data=pdf_data,
                            file_name=f"{reference_number}_cut_list.pdf",
                            mime="application/pdf",
                            use_container_width=True
                        )
            else:
                st.warning("No kickplate items (starting with 'KP') found in the CSV")
    
    # ========== TAB 2: Manual Entry ==========
    with tab2:
        st.subheader("Manual Entry")
        st.info("Add kickplates manually using the form below")
        
        with st.form("add_item_form"):
            col1, col2, col3 = st.columns(3)
            with col1:
                width = st.number_input("Width (mm)", min_value=1, value=300, step=10, key="manual_width")
            with col2:
                height = st.number_input("Height (mm)", min_value=1, value=300, step=10, key="manual_height")
            with col3:
                qty = st.number_input("Quantity", min_value=1, value=1, key="manual_qty")
            
            col4, col5 = st.columns(2)
            with col4:
                material = st.text_input("Material Code", value="SSS", help="Default: SSS (Stainless Steel)", key="manual_material").upper()
            with col5:
                description = st.text_input("Description (optional)", value="", placeholder="e.g., Kitchen kickplate", key="manual_desc")
            
            submitted = st.form_submit_button("Add Item", use_container_width=True)
            if submitted:
                code = f"KP{width}{height}{material}"
                if not description:
                    description = f"{width}×{height}mm {material}"
                
                st.session_state.manual_items.append({
                    'code': code,
                    'width': width,
                    'height': height,
                    'material': material,
                    'description': description,
                    'qty': qty
                })
                st.success(f"Added {code} ({width}×{height}mm) × {qty}")
                st.rerun()
        
        if st.session_state.manual_items:
            st.write("### Current items:")
            display_items = []
            for item in st.session_state.manual_items:
                display_items.append({
                    'Part Code': item['code'],
                    'Width (mm)': item['width'],
                    'Height (mm)': item['height'],
                    'Material': item['material'],
                    'Description': item['description'],
                    'Quantity': item['qty']
                })
            items_df = pd.DataFrame(display_items)
            st.dataframe(items_df, use_container_width=True)
            
            col1, col2 = st.columns(2)
            with col1:
                if st.button("Clear All", use_container_width=True, key="manual_clear"):
                    st.session_state.manual_items = []
                    st.session_state.manual_sheets = None
                    st.rerun()
            with col2:
                if st.button("🚀 Generate Cut List", type="primary", use_container_width=True, key="manual_generate"):
                    nester = KickplateNester(stock_width, stock_height, kerf_width, grain_direction)
                    pieces = [Piece(
                        code=item['code'],
                        description=item['description'],
                        width=item['width'],
                        height=item['height'],
                        material=item['material'],
                        qty=item['qty']
                    ) for item in st.session_state.manual_items]
                    
                    with st.spinner("Optimizing layout..."):
                        st.session_state.manual_sheets = nester.nest_pieces(pieces)
                    st.rerun()
        
        if st.session_state.manual_sheets:
            sheets = st.session_state.manual_sheets
            st.success(f"✅ Generated cut list with {len(sheets)} sheet(s)")
            
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Sheets Required", len(sheets))
            with col2:
                total_pieces = sum(len(s.placements) for s in sheets)
                st.metric("Total Pieces", total_pieces)
            with col3:
                avg_efficiency = sum(100 - (s.waste_area / (stock_width * stock_height) * 100) for s in sheets) / len(sheets)
                st.metric("Avg Efficiency", f"{avg_efficiency:.1f}%")
            
            for sheet in sheets:
                with st.expander(f"📋 Sheet {sheet.id + 1} - {len(sheet.placements)} pieces", expanded=True):
                    fig = create_sheet_visualization(sheet, stock_width, stock_height, grain_direction)
                    st.plotly_chart(fig, use_container_width=True)
                    
                    st.subheader("Cutting Instructions")
                    cut_data = []
                    for i, p in enumerate(sheet.placements):
                        cut_data.append({
                            'Piece #': i + 1,
                            'Part Code': p.code,
                            'Description': p.description,
                            'X (mm)': p.x,
                            'Y (mm)': p.y,
                            'Width (mm)': p.width,
                            'Height (mm)': p.height,
                            'Rotated': '↻ Yes' if p.rotated else 'No'
                        })
                    st.dataframe(pd.DataFrame(cut_data), use_container_width=True)
            
            cut_list_data = []
            for sheet in sheets:
                for i, p in enumerate(sheet.placements):
                    cut_list_data.append({
                        'Sheet': sheet.id + 1,
                        'Piece #': i + 1,
                        'Part Code': p.code,
                        'Description': p.description,
                        'X Position': p.x,
                        'Y Position': p.y,
                        'Width': p.width,
                        'Height': p.height,
                        'Rotated': 'Yes' if p.rotated else 'No'
                    })
            
            csv_data = pd.DataFrame(cut_list_data).to_csv(index=False)
            
            col1, col2 = st.columns(2)
            with col1:
                st.download_button(
                    label="📥 Download CSV",
                    data=csv_data,
                    file_name="manual_cut_list.csv",
                    mime="text/csv",
                    use_container_width=True
                )
            with col2:
                pdf_data = generate_pdf_cutlist(sheets, stock_width, stock_height, grain_direction,
                                               reference_number=None)
                st.download_button(
                    label="📄 Download PDF",
                    data=pdf_data,
                    file_name="manual_cut_list.pdf",
                    mime="application/pdf",
                    use_container_width=True
                )
    
    # ========== TAB 3: Door Labels ==========
    with tab3:
        st.subheader("🏷️ Door Label Generator")
        
        parser = DoorLabelParser()
        
        # File upload section
        upload_col1, upload_col2 = st.columns([2, 1])
        
        with upload_col1:
            file_type = st.selectbox(
                "Select File Type",
                ["Finishing Sheet (PDF)", "Order CSV", "Door Schedule (Excel)", "Manual Entry"],
                key="label_file_type"
            )
            
            if file_type != "Manual Entry":
                uploaded_file = st.file_uploader(
                    f"Upload {file_type.split('(')[1][:-1]} file",
                    type=['pdf', 'csv', 'xlsx', 'xls'],
                    key="label_uploader"
                )
        
        with upload_col2:
            st.write("### Project Info")
            
            # Auto-detect from PDF
            if uploaded_file and file_type == "Finishing Sheet (PDF)":
                if st.button("🔍 Auto-detect Project Info", key="auto_detect"):
                    with st.spinner("Extracting project info..."):
                        try:
                            # Extract project info from PDF
                            project_info = parser._extract_project_info_from_pdf(uploaded_file)
                            st.session_state.parsed_project_info = project_info
                            
                            # Also update main project info
                            st.session_state.project_info = project_info
                            st.success(f"✅ Detected: {project_info['project_code']} - {project_info['project_name']}")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Could not auto-detect project info: {e}")
            
            # Show current project info (auto-detected or manual)
            if st.session_state.parsed_project_info:
                # Use auto-detected info
                project_info = st.session_state.parsed_project_info
                st.info(f"**Auto-detected:**\n"
                       f"Code: `{project_info['project_code']}`\n"
                       f"Name: `{project_info['project_name']}`")
            else:
                # Allow manual override
                use_sidebar_info = st.checkbox("Use sidebar project info", value=True)
                
                if not use_sidebar_info:
                    project_code_input = st.text_input("Project Code", key="label_project_code")
                    project_name_input = st.text_input("Project Name", key="label_project_name")
                    project_info = {
                        'project_code': project_code_input,
                        'project_name': project_name_input
                    }
                else:
                    project_info = st.session_state.project_info
        
        # Parse button - UPDATED TO AUTO-FILL PROJECT INFO
        if file_type != "Manual Entry" and uploaded_file:
            col1, col2 = st.columns(2)
            with col1:
                if st.button("📥 Parse File & Generate Labels", type="primary", use_container_width=True):
                    with st.spinner("Parsing file..."):
                        try:
                            if file_type == "Finishing Sheet (PDF)":
                                # First extract project info if not already done
                                if st.session_state.parsed_project_info is None:
                                    extracted_info = parser._extract_project_info_from_pdf(uploaded_file)
                                    st.session_state.parsed_project_info = extracted_info
                                    project_info = extracted_info
                                    st.success(f"📋 Auto-detected project: {extracted_info['project_code']}")
                                
                                labels = parser.parse_pdf_fixed(uploaded_file, project_info)
                            elif file_type == "Order CSV":
                                labels = parser.parse_csv(uploaded_file, project_info)
                            elif file_type == "Door Schedule (Excel)":
                                labels = parser.parse_excel(uploaded_file, project_info)
                            
                            st.session_state.door_labels = labels
                            st.success(f"✅ Parsed {len(labels)} door labels")
                            
                            # Show sample of parsed data
                            if labels:
                                st.write("### Sample of Parsed Labels (First 10):")
                                sample_data = []
                                for i, label in enumerate(labels[:10]):
                                    sample_data.append({
                                        'Door': label.door_number,
                                        'Area': label.area,
                                        'Size': f"{label.width}×{label.height}mm",
                                        'Material': label.material,
                                        'Project': label.project_code
                                    })
                                sample_df = pd.DataFrame(sample_data)
                                st.dataframe(sample_df)
                                
                        except Exception as e:
                            st.error(f"Error parsing file: {str(e)}")
            
            with col2:
                if st.button("🔍 Quick Debug", key="quick_debug"):
                    if uploaded_file and file_type == "Finishing Sheet (PDF)":
                        with pdfplumber.open(uploaded_file) as pdf:
                            first_page = pdf.pages[0]
                            text = first_page.extract_text()
                            
                            st.write("### First 1000 chars of PDF:")
                            st.text(text[:1000])
                            
                            # Also show project info extraction attempt
                            project_info = parser._extract_project_info_from_pdf(uploaded_file)
                            st.write(f"### Auto-detected Project Info:")
                            st.write(f"Code: `{project_info['project_code']}`")
                            st.write(f"Name: `{project_info['project_name']}`")
        
        # Manual entry for labels
        elif file_type == "Manual Entry":
            st.write("### Manual Door Entry")
            
            with st.form("manual_door_form"):
                col1, col2 = st.columns(2)
                with col1:
                    door_number = st.text_input("Door Number", placeholder="e.g., D.26-K or W.15")
                    area = st.text_input("Area/Location", placeholder="e.g., PE Office, WC 2")
                with col2:
                    kickplate_code = st.text_input("Kickplate Code", placeholder="e.g., KP800300SSS")
                    width = st.number_input("Width (mm)", min_value=1, value=800, step=10)
                    height = st.number_input("Height (mm)", min_value=1, value=300, step=10)
                
                add_door = st.form_submit_button("Add Door", use_container_width=True)
                
                if add_door and door_number and kickplate_code:
                    match = re.match(parser.kickplate_pattern, kickplate_code)
                    if match:
                        material = match.group(3)
                        label = DoorLabel(
                            door_number=door_number,
                            area=area or "Unknown",
                            kickplate_code=kickplate_code,
                            width=width,
                            height=height,
                            project_code=project_info['project_code'],
                            project_name=project_info['project_name'],
                            material=material,
                            quantity=1
                        )
                        st.session_state.door_labels.append(label)
                        st.success(f"Added door {door_number}")
                        st.rerun()
                    else:
                        st.error("Invalid kickplate code format. Use format like KP800300SSS")
        
        # Display and edit labels
        if st.session_state.door_labels:
            st.write(f"### 📋 {len(st.session_state.door_labels)} Door Labels")
            
            # Edit/delete interface
            edit_col1, edit_col2 = st.columns(2)
            
            with edit_col1:
                if st.button("🔄 Refresh Display", use_container_width=True):
                    st.rerun()
            
            with edit_col2:
                if st.button("🗑️ Clear All Labels", use_container_width=True):
                    st.session_state.door_labels = []
                    st.session_state.label_pieces = None
                    st.session_state.label_sheets = None
                    st.session_state.parsed_project_info = None
                    st.rerun()
            
            # Display labels in an editable dataframe
            label_data = []
            for i, label in enumerate(st.session_state.door_labels):
                label_data.append({
                    'ID': i + 1,
                    'Door Number': label.door_number,
                    'Area': label.area,
                    'Kickplate': f"{label.width}×{label.height}mm",
                    'Project': label.project_code
                })
            
            if label_data:
                df_labels = pd.DataFrame(label_data)
                edited_df = st.data_editor(
                    df_labels,
                    use_container_width=True,
                    num_rows="dynamic",
                    key="label_editor",
                    column_config={
                        "Door Number": st.column_config.TextColumn(
                            "Door Number",
                            help="Door identifier like D.26-K or W.15"
                        ),
                        "Area": st.column_config.TextColumn(
                            "Area/Location",
                            help="Room or area name"
                        ),
                        "Kickplate": st.column_config.TextColumn(
                            "Kickplate Size",
                            help="Width × Height in mm"
                        )
                    }
                )
                
                # Update labels from edited dataframe
                if not edited_df.equals(df_labels):
                    for idx, row in edited_df.iterrows():
                        if idx < len(st.session_state.door_labels):
                            st.session_state.door_labels[idx].door_number = row['Door Number']
                            st.session_state.door_labels[idx].area = row['Area']
            
            # Label preview
            st.write("### 🔍 Label Preview")
            preview_cols = st.columns(3)
            
            for i, (label, col) in enumerate(zip(st.session_state.door_labels[:3], preview_cols)):
                with col:
                    with st.container(border=True):
                        st.markdown(f"**Door {i+1} Preview:**")
                        st.code(label.label_text(include_project=include_project), language=None)
            
            # 🔗 QUICK INTEGRATION BUTTON
            st.write("### 🔗 Integration with Nesting Optimizer")
            
            col1, col2 = st.columns(2)
            
            with col1:
                if st.button("🚀 Generate Cut List from These Labels", type="primary", use_container_width=True):
                    # Extract pieces from labels
                    pieces = extract_pieces_from_labels(st.session_state.door_labels)
                    
                    # Create nester with current sidebar settings
                    nester = KickplateNester(
                        stock_width=stock_width,
                        stock_height=stock_height,
                        kerf_width=kerf_width,
                        grain_direction=grain_direction
                    )
                    
                    with st.spinner("Optimizing cut layout..."):
                        sheets = nester.nest_pieces(pieces)
                    
                    # Store in session state
                    st.session_state.label_sheets = sheets
                    st.session_state.label_pieces = pieces
                    st.success(f"✅ Generated cut list with {len(sheets)} sheet(s) from {len(pieces)} unique kickplate sizes")
            
            with col2:
                if st.button("📊 Show Kickplate Summary", use_container_width=True):
                    pieces = extract_pieces_from_labels(st.session_state.door_labels)
                    
                    st.write("### Kickplate Summary for Nesting")
                    summary_data = []
                    total_pieces = 0
                    
                    for piece in pieces:
                        summary_data.append({
                            'Size': f"{piece.width}×{piece.height}mm",
                            'Material': piece.material,
                            'Quantity': piece.qty,
                            'Code': piece.code
                        })
                        total_pieces += piece.qty
                    
                    summary_df = pd.DataFrame(summary_data)
                    st.dataframe(summary_df, use_container_width=True)
                    
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("Unique Sizes", len(pieces))
                    with col2:
                        st.metric("Total Pieces", total_pieces)
                    with col3:
                        total_area = sum(p.width * p.height * p.qty for p in pieces) / 1000000  # m²
                        st.metric("Total Area", f"{total_area:.2f} m²")
            
            # Show cut list results if generated
            if st.session_state.label_sheets:
                sheets = st.session_state.label_sheets
                pieces = st.session_state.label_pieces
                
                st.write("### 📋 Cut List Results")
                
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric("Sheets Required", len(sheets))
                with col2:
                    total_pieces = sum(len(s.placements) for s in sheets)
                    st.metric("Total Pieces", total_pieces)
                with col3:
                    avg_efficiency = sum(100 - (s.waste_area / (stock_width * stock_height) * 100) for s in sheets) / len(sheets)
                    st.metric("Avg Efficiency", f"{avg_efficiency:.1f}%")
                
                # Visualize first sheet
                if sheets:
                    with st.expander("📋 View First Sheet Layout", expanded=True):
                        fig = create_sheet_visualization(sheets[0], stock_width, stock_height, grain_direction)
                        st.plotly_chart(fig, use_container_width=True)
                
                # Download options for cut list
                st.write("#### 📥 Download Cut List")
                
                cut_list_data = []
                for sheet in sheets:
                    for i, p in enumerate(sheet.placements):
                        cut_list_data.append({
                            'Sheet': sheet.id + 1,
                            'Piece #': i + 1,
                            'Part Code': p.code,
                            'Description': p.description,
                            'X Position': p.x,
                            'Y Position': p.y,
                            'Width': p.width,
                            'Height': p.height,
                            'Rotated': 'Yes' if p.rotated else 'No'
                        })
                
                cut_list_csv = pd.DataFrame(cut_list_data).to_csv(index=False)
                cut_list_pdf = generate_pdf_cutlist(sheets, stock_width, stock_height, grain_direction, 
                                                   reference_number=project_info['project_code'])
                
                col1, col2 = st.columns(2)
                with col1:
                    st.download_button(
                        label="🔪 Cut List CSV",
                        data=cut_list_csv,
                        file_name=f"{project_info['project_code']}_cut_list.csv",
                        mime="text/csv",
                        use_container_width=True
                    )
                with col2:
                    st.download_button(
                        label="📄 Cut List PDF",
                        data=cut_list_pdf,
                        file_name=f"{project_info['project_code']}_cut_list.pdf",
                        mime="application/pdf",
                        use_container_width=True
                    )
            
            # Download options for labels
            st.write("### 📥 Download Label Files")
            
            if st.session_state.door_labels:
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    # CSV export
                    csv_data = pd.DataFrame([label.to_dict() for label in st.session_state.door_labels])
                    csv_bytes = csv_data.to_csv(index=False).encode('utf-8')
                    
                    st.download_button(
                        label="📄 Door Labels CSV",
                        data=csv_bytes,
                        file_name=f"door_labels_{project_info['project_code']}.csv",
                        mime="text/csv",
                        use_container_width=True
                    )
                
                with col2:
                    # Label PDF
                    label_pdf = generate_label_pdf(
                        st.session_state.door_labels,
                        label_size=(label_width, label_height),
                        labels_per_page=(3, 8),
                        include_project=include_project
                    )
                    
                    st.download_button(
                        label="🏷️ Door Labels PDF",
                        data=label_pdf,
                        file_name=f"door_labels_{project_info['project_code']}.pdf",
                        mime="application/pdf",
                        use_container_width=True
                    )
                
                with col3:
                    # Summary report
                    summary_data = []
                    for label in st.session_state.door_labels:
                        summary_data.append({
                            'Door Number': label.door_number,
                            'Area': label.area,
                            'Width (mm)': label.width,
                            'Height (mm)': label.height,
                            'Material': label.material
                        })
                    
                    summary_df = pd.DataFrame(summary_data)
                    summary_csv = summary_df.to_csv(index=False).encode('utf-8')
                    
                    st.download_button(
                        label="📊 Door Summary CSV",
                        data=summary_csv,
                        file_name=f"door_summary_{project_info['project_code']}.csv",
                        mime="text/csv",
                        use_container_width=True
                    )
        else:
            st.info("👆 Upload a finishing sheet PDF or enter doors manually to generate labels")
    
    # ========== TAB 4: Debug Tools ==========
    with tab4:
        st.subheader("🔧 Debug Tools")
        
        st.write("### PDF Parser Debug")
        
        debug_file = st.file_uploader("Upload PDF for debugging", type=['pdf'], key="debug_uploader")
        
        if debug_file:
            if st.button("Run Deep Debug Analysis", use_container_width=True):
                with pdfplumber.open(debug_file) as pdf:
                    st.write("### Page 1 Analysis")
                    
                    # Show raw text
                    first_page = pdf.pages[0]
                    text = first_page.extract_text()
                    
                    st.write("**Raw Text (First 3000 chars):**")
                    st.text(text[:3000])
                    
                    st.write("**Text Lines:**")
                    lines = text.split('\n')
                    for i, line in enumerate(lines[:50]):  # First 50 lines
                        st.write(f"{i+1}: `{line}`")
        
        st.write("### Session State Debug")
        if st.button("Show Session State", key="show_session"):
            st.write("**door_labels:**", len(st.session_state.get('door_labels', [])))
            st.write("**project_info:**", st.session_state.get('project_info', {}))
            st.write("**parsed_project_info:**", st.session_state.get('parsed_project_info', {}))
            st.write("**manual_items:**", len(st.session_state.get('manual_items', [])))
            st.write("**label_pieces:**", st.session_state.get('label_pieces', None) is not None)
            st.write("**label_sheets:**", st.session_state.get('label_sheets', None) is not None)

if __name__ == "__main__":
    main()